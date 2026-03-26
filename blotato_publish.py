"""
blotato_publish.py — Publica posts desde BLOTATO_POSTS.xlsx a través de la API de Blotato.

Uso:
  python blotato_publish.py              → procesa todas las filas "Pendiente"
  python blotato_publish.py --dry-run    → muestra lo que haría sin publicar
  python blotato_publish.py --row 5      → procesa sólo la fila 5
  python blotato_publish.py --status     → revisa el estado de posts ya enviados

Requisitos:
  pip install openpyxl requests
"""

import os, sys, json, base64, argparse, traceback
from datetime import datetime, timezone
from pathlib import Path

import requests
import openpyxl
from openpyxl import load_workbook

# ── Configuración ─────────────────────────────────────────────────────────────
API_KEY    = "blt_PXDsPBb6/fGzCMNaqRLMXVnca/HC+A5JYqP8kIYkUHo="
BASE_URL   = "https://backend.blotato.com/v2"
SHEET_PATH = r"C:\Users\Usuario\Desktop\BLOTATO_POSTS.xlsx"
LOG_PATH   = r"C:\Users\Usuario\Desktop\BLOTATO_LOG.txt"

HEADERS = {"blotato-api-key": API_KEY, "Content-Type": "application/json"}

# ── Logging ───────────────────────────────────────────────────────────────────
def log(msg):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line)
    with open(LOG_PATH, "a", encoding="utf-8") as f:
        f.write(line + "\n")

# ── Helpers de columnas ───────────────────────────────────────────────────────
def col_map(ws):
    """Devuelve dict {nombre_columna: indice_1based} desde la fila 1."""
    return {ws.cell(row=1, column=i).value: i
            for i in range(1, ws.max_column + 1)
            if ws.cell(row=1, column=i).value}

def get(row, cols, name, default=None):
    ci = cols.get(name)
    if ci is None:
        return default
    val = row[ci - 1].value
    return val if val is not None else default

def set_cell(ws, row_num, cols, name, value):
    ci = cols.get(name)
    if ci:
        ws.cell(row=row_num, column=ci).value = value

# ── Upload de media ────────────────────────────────────────────────────────────
def upload_media(path_or_url: str) -> str:
    """
    Si es URL pública → la devuelve directamente (Blotato la descarga sola).
    Si es ruta local  → la codifica en base64 y la sube al endpoint /media.
    Devuelve la URL hosteada en Blotato.
    """
    if path_or_url.startswith("http://") or path_or_url.startswith("https://"):
        return path_or_url

    # Es ruta local
    p = Path(path_or_url)
    if not p.exists():
        raise FileNotFoundError(f"Archivo no encontrado: {path_or_url}")

    size_mb = p.stat().st_size / (1024 * 1024)
    log(f"  Subiendo media local ({size_mb:.1f} MB): {p.name}")

    suffix = p.suffix.lower()
    mime_map = {
        ".mp4": "video/mp4", ".mov": "video/quicktime",
        ".avi": "video/x-msvideo", ".mkv": "video/x-matroska",
        ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
        ".png": "image/png",  ".gif": "image/gif",
        ".webp": "image/webp",
    }
    mime = mime_map.get(suffix, "application/octet-stream")

    with open(p, "rb") as f:
        b64 = base64.b64encode(f.read()).decode("utf-8")
    data_uri = f"data:{mime};base64,{b64}"

    resp = requests.post(
        f"{BASE_URL}/media",
        headers=HEADERS,
        json={"url": data_uri},
        timeout=120,
    )
    resp.raise_for_status()
    hosted_url = resp.json()["url"]
    log(f"  Media subida: {hosted_url}")
    return hosted_url

# ── Construcción del body por plataforma ──────────────────────────────────────
def build_body(row, cols):
    platform   = str(get(row, cols, "Plataforma", "")).strip().lower()
    account_id = str(get(row, cols, "Account_ID", "")).strip()
    texto      = str(get(row, cols, "Texto_Copy", "") or "")
    hashtags   = str(get(row, cols, "Hashtags",   "") or "")
    video_url  = str(get(row, cols, "Video_URL",  "") or "").strip()
    media_extra= str(get(row, cols, "Media_Extra","") or "").strip()
    fecha      = get(row, cols, "Fecha")
    hora       = str(get(row, cols, "Hora", "") or "").strip()
    zona       = str(get(row, cols, "Zona_Horaria", "America/Mexico_City") or "America/Mexico_City").strip()
    slot_libre = str(get(row, cols, "Usar_Slot_Libre", "FALSE")).strip().upper() == "TRUE"

    if not platform:
        raise ValueError("Columna 'Plataforma' vacía")
    if not account_id:
        raise ValueError("Columna 'Account_ID' vacía")

    # ── Texto final = copy + hashtags ──────────────────────────────────────
    full_text = texto.strip()
    if hashtags.strip():
        full_text = f"{full_text}\n\n{hashtags.strip()}"

    # ── Media URLs ─────────────────────────────────────────────────────────
    media_urls = []
    if video_url:
        media_urls.append(upload_media(video_url))
    for extra in media_extra.split(","):
        extra = extra.strip()
        if extra:
            media_urls.append(upload_media(extra))

    # ── Target por plataforma ──────────────────────────────────────────────
    target = {"targetType": platform}

    if platform == "instagram":
        ig_tipo     = str(get(row, cols, "IG_Tipo",        "") or "").strip().lower()
        ig_alt      = get(row, cols, "IG_AltText")
        ig_cover    = str(get(row, cols, "IG_Portada_URL", "") or "").strip()
        ig_feed     = str(get(row, cols, "IG_ShareToFeed", "FALSE")).strip().upper() == "TRUE"
        ig_audio    = get(row, cols, "IG_AudioName")
        if ig_tipo in ("reel", "story"):
            target["mediaType"] = ig_tipo
        if ig_alt:
            target["altText"] = ig_alt
        if ig_cover:
            target["coverImageUrl"] = upload_media(ig_cover)
        target["shareToFeed"] = ig_feed
        if ig_audio:
            target["audioName"] = ig_audio

    elif platform == "tiktok":
        privacy  = str(get(row, cols, "TT_Privacidad",  "PUBLIC_TO_EVERYONE") or "PUBLIC_TO_EVERYONE").strip()
        target["privacyLevel"]       = privacy
        target["disabledComments"]   = str(get(row, cols, "TT_Comentarios","FALSE")).strip().upper() == "TRUE"
        target["disabledDuet"]       = str(get(row, cols, "TT_Dueto",      "FALSE")).strip().upper() == "TRUE"
        target["disabledStitch"]     = str(get(row, cols, "TT_Stitch",     "FALSE")).strip().upper() == "TRUE"
        target["isBrandedContent"]   = str(get(row, cols, "TT_Branded",    "FALSE")).strip().upper() == "TRUE"
        target["isYourBrand"]        = str(get(row, cols, "TT_TuMarca",    "FALSE")).strip().upper() == "TRUE"
        target["isAiGenerated"]      = str(get(row, cols, "TT_EsIA",       "TRUE" )).strip().upper() == "TRUE"

    elif platform == "youtube":
        yt_title   = str(get(row, cols, "YT_Titulo",    "") or "").strip()
        yt_privacy = str(get(row, cols, "YT_Privacidad","public") or "public").strip().lower()
        yt_notify  = str(get(row, cols, "YT_Notificar", "FALSE")).strip().upper() == "TRUE"
        yt_kids    = str(get(row, cols, "YT_ParaNinos",  "FALSE")).strip().upper() == "TRUE"
        yt_ai      = str(get(row, cols, "YT_EsIA",       "TRUE" )).strip().upper() == "TRUE"
        yt_thumb   = str(get(row, cols, "YT_Miniatura",  "") or "").strip()
        if not yt_title:
            raise ValueError("YouTube requiere 'YT_Titulo'")
        target["title"]                 = yt_title
        target["privacyStatus"]         = yt_privacy
        target["shouldNotifySubscribers"] = yt_notify
        target["isMadeForKids"]         = yt_kids
        target["containsSyntheticMedia"]= yt_ai
        if yt_thumb:
            target["thumbnailUrl"] = upload_media(yt_thumb)

    elif platform == "facebook":
        fb_page    = str(get(row, cols, "Account_ID", "")).strip()  # pageId = accountId para FB
        fb_media   = str(get(row, cols, "FB_TipoMedia","") or "").strip().lower()
        fb_link    = str(get(row, cols, "FB_Link",      "") or "").strip()
        target["pageId"] = fb_page
        if fb_media in ("video", "reel"):
            target["mediaType"] = fb_media
        if fb_link:
            target["link"] = fb_link

    elif platform == "threads":
        pass  # sólo targetType

    # ── Programación ──────────────────────────────────────────────────────
    body = {
        "post": {
            "accountId": account_id,
            "content": {
                "text":        full_text,
                "mediaUrls":   media_urls,
                "platform":    platform,
                "additionalPosts": [],
            },
            "target": target,
        }
    }

    if slot_libre:
        body["useNextFreeSlot"] = True
    elif fecha and hora:
        # Combina fecha + hora con timezone
        if hasattr(fecha, "strftime"):
            fecha_str = fecha.strftime("%Y-%m-%d")
        else:
            fecha_str = str(fecha).split(" ")[0].split("T")[0]
        hora_str = hora.strip()
        # Offset básico para zonas comunes
        tz_offsets = {
            "America/Mexico_City": "-06:00",
            "America/New_York":    "-05:00",
            "America/Chicago":     "-06:00",
            "America/Los_Angeles": "-08:00",
            "UTC":                 "+00:00",
            "Europe/Madrid":       "+01:00",
        }
        offset = tz_offsets.get(zona, "-06:00")
        body["scheduledTime"] = f"{fecha_str}T{hora_str}:00{offset}"
    # Si no hay fecha ni slot → publicación inmediata

    return body

# ── Publicar un post ──────────────────────────────────────────────────────────
def publish(body: dict) -> dict:
    resp = requests.post(
        f"{BASE_URL}/posts",
        headers=HEADERS,
        json=body,
        timeout=60,
    )
    try:
        data = resp.json()
    except Exception:
        data = {"raw": resp.text}
    if not resp.ok:
        raise RuntimeError(f"HTTP {resp.status_code}: {data}")
    return data

# ── Revisar estado de un post ─────────────────────────────────────────────────
def check_status(submission_id: str) -> dict:
    resp = requests.get(
        f"{BASE_URL}/posts/{submission_id}",
        headers=HEADERS,
        timeout=20,
    )
    try:
        return resp.json()
    except Exception:
        return {"raw": resp.text}

# ── Procesamiento principal ───────────────────────────────────────────────────
def run(dry_run=False, only_row=None, check_mode=False):
    if not Path(SHEET_PATH).exists():
        log(f"ERROR: No se encontro el archivo: {SHEET_PATH}")
        sys.exit(1)

    wb = load_workbook(SHEET_PATH)
    ws = wb["POSTS"]
    cols = col_map(ws)

    if check_mode:
        log("=== Revisando estado de posts enviados ===")
        updated = 0
        for row_num in range(3, ws.max_row + 1):
            row = list(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=False))[0]
            estado = str(get(row, cols, "Estado", "")).strip()
            sub_id = str(get(row, cols, "Post_Submission_ID", "") or "").strip()
            if estado in ("Enviado", "Pendiente_Check") and sub_id:
                try:
                    status = check_status(sub_id)
                    nuevo_estado = status.get("status", "Desconocido")
                    set_cell(ws, row_num, cols, "Estado_API", nuevo_estado)
                    set_cell(ws, row_num, cols, "Estado", "Publicado" if nuevo_estado == "published" else estado)
                    log(f"  Fila {row_num}: {sub_id} → {nuevo_estado}")
                    updated += 1
                except Exception as e:
                    log(f"  Fila {row_num}: ERROR al revisar — {e}")
        wb.save(SHEET_PATH)
        log(f"Check completo. {updated} filas actualizadas.")
        return

    log("=== Blotato Publisher ===")
    log(f"Modo: {'DRY RUN (sin publicar)' if dry_run else 'PUBLICACION REAL'}")

    total = ok = errors = skipped = 0

    for row_num in range(3, ws.max_row + 1):
        row = list(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=False))[0]

        estado = str(get(row, cols, "Estado", "")).strip()
        if not any(get(row, cols, c) for c in ["Plataforma", "Texto_Copy", "Video_URL"]):
            break  # Fin de datos reales

        if only_row and row_num != only_row:
            continue

        if estado != "Pendiente":
            skipped += 1
            continue

        total += 1
        plataforma = get(row, cols, "Plataforma", "?")
        cuenta     = get(row, cols, "Cuenta_Nombre", get(row, cols, "Account_ID", "?"))
        log(f"\n[Fila {row_num}] {plataforma} | {cuenta}")

        try:
            body = build_body(row, cols)

            if dry_run:
                log(f"  DRY RUN body: {json.dumps(body, ensure_ascii=False, indent=2)[:500]}...")
                ok += 1
                continue

            result = publish(body)
            sub_id = result.get("postSubmissionId", "")
            log(f"  OK → postSubmissionId: {sub_id}")

            set_cell(ws, row_num, cols, "Post_Submission_ID", sub_id)
            set_cell(ws, row_num, cols, "Estado",             "Enviado")
            set_cell(ws, row_num, cols, "Estado_API",         "submitted")
            set_cell(ws, row_num, cols, "Fecha_Publicado",    datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            ok += 1

        except Exception as e:
            err_msg = str(e)
            log(f"  ERROR: {err_msg}")
            traceback.print_exc()
            set_cell(ws, row_num, cols, "Estado", "Error")
            set_cell(ws, row_num, cols, "Notas",  err_msg[:500])
            errors += 1

        finally:
            if not dry_run:
                wb.save(SHEET_PATH)

    log(f"\n=== Resumen ===")
    log(f"  Procesados : {total}")
    log(f"  OK         : {ok}")
    log(f"  Errores    : {errors}")
    log(f"  Saltados   : {skipped}  (no estaban en Pendiente)")
    if not dry_run:
        log(f"  Archivo guardado: {SHEET_PATH}")
        log(f"  Log guardado   : {LOG_PATH}")


# ── CLI ───────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Blotato Publisher")
    parser.add_argument("--dry-run",  action="store_true", help="Simula sin publicar")
    parser.add_argument("--row",      type=int,            help="Procesar solo la fila N")
    parser.add_argument("--status",   action="store_true", help="Revisar estado de posts enviados")
    args = parser.parse_args()

    # Instalar requests si no está
    try:
        import requests
    except ImportError:
        import subprocess
        subprocess.run([sys.executable, "-m", "pip", "install", "requests", "-q"])
        import requests

    run(dry_run=args.dry_run, only_row=args.row, check_mode=args.status)
